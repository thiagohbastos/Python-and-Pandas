# %%
##### Pacotes #####
import os
import numpy as np
import pandas as pd
from datetime import datetime, date, timedelta

from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, numbers, PatternFill, Protection

import sqlalchemy
from sqlalchemy.engine import URL

import warnings
warnings.filterwarnings('ignore')



# %%
##### DEFININDO FUNÇÕES #####
def paint(cell_range, color):
    if color.lower() == 'azul':
        fill_color = '4f81bd'
        font_color = 'ffffff'
    elif color.lower() == 'marrom':
        fill_color = '833c0c'
        font_color = 'ffffff'
    elif color.lower() == 'amarelo':
        fill_color = 'ffff00'
        font_color = '000000'
    elif color.lower() == 'azul claro':
        fill_color = 'dce6f1'
        font_color = '000000'
    else:
        fill_color = 'ffffff'
        font_color = '000000'

    for row in cell_range:
        for cell in row:
            cell.fill = PatternFill(start_color=fill_color,
                                    end_color=fill_color,
                                    fill_type='solid')
            cell.font = Font(bold=True, color=font_color)

def money_format(col):
    for cell in col:
        cell.number_format = '#,##0.00'

def os_format(col):
    for cell in col:
        cell.number_format = '0'

def add_validation(ws, cells, options):
    dv = DataValidation(type="list", formula1=options, allow_blank=True)

    dv.error ='Selecione uma opção da lista.'
    dv.errorTitle = 'Valor inválido'
    dv.prompt = 'Selecione uma opção:'
    #dv.promptTitle = 'Opções'

    dv.showErrorMessage = True
    dv.showInputMessage = True

    ws.add_data_validation(dv)
    dv.add(cells)

def size_and_alignment(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 5)
        sheet.column_dimensions[column].width = adjusted_width

def block_edit(ws):
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.formatCells = False
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.password = "fatBMB"
    ws.protection.enable()

def formatar_arquivo(file, cod_imposto):
    wb = load_workbook(file)

    ### BASE
    ws = wb['BASE']
    ws.auto_filter.ref = ws.dimensions
    paint(ws["A1":"J1"], 'azul')
    paint(ws["K1":"T1"], 'marrom')
    money_format(ws['I:I'])
    add_validation(ws, "L2:L20000", tip_options)
    add_validation(ws, "S2:S20000", obs_options)
    os_format(ws['J:J'])
    size_and_alignment(ws)
    block_edit(ws)
    for col in ["K","L","M","N","O","P","Q","R","S","T"]:
        for cell in ws[col]:
            cell.protection = Protection(locked=False)

    ### INCLUIR BASE
    ws = wb['INCLUIR BASE']
    ws.auto_filter.ref = ws.dimensions
    paint(ws["A1":"J1"], 'azul')
    paint(ws["K1":"S1"], 'marrom')
    money_format(ws['I:I'])
    add_validation(ws, "K2:K5000", tip_options)
    add_validation(ws, "R2:R5000", obs_options)
    size_and_alignment(ws)

    ### LEGENDA
    ws = wb['LEGENDA']
    paint(ws["A1":"B1"], 'amarelo')
    paint(ws["D1":"E1"], 'amarelo')
    size_and_alignment(ws)
    block_edit(ws)

    if cod_imposto == 1:
        ### CUSTODIA
        ws = wb['CUSTODIA']
        ws.auto_filter.ref = ws.dimensions
        paint(ws["A1":"F1"], 'azul')
        paint(ws["G1":"K1"], 'marrom')
        money_format(ws['F:F'])
        size_and_alignment(ws)
        block_edit(ws)
        for col in ["G","H","I","J","K"]:
            for cell in ws[col]:
                cell.protection = Protection(locked=False)

        ### INCLUIR CUSTODIA
        ws = wb['INCLUIR CUSTODIA']
        ws.auto_filter.ref = ws.dimensions
        paint(ws["A1":"F1"], 'azul')
        paint(ws["G1":"K1"], 'marrom')
        money_format(ws['F:F'])
        size_and_alignment(ws)

    wb.save(file)

def gerar_base(dta_inicial_padrao, resp_dta_fim, resp_transp, resp_imposto):
    ### BASE
    query_base = f"""
    DECLARE @dta_corte AS DATE 
    DECLARE @dta_limite AS DATE 
    DECLARE @transp AS VARCHAR(50)
    DECLARE @imposto AS INT

    SET @dta_corte = '{dta_inicial_padrao}'
    SET @dta_limite = '{resp_dta_fim}'
    SET @transp = '{resp_transp}'
    SET @imposto = {resp_imposto}

    SELECT 
         A.COD_SVC
        ,A.SVC
        ,A.DATA_MOVIMENTO AS DATA
        ,A.COD_CEN
        ,A.DES_CEN
        ,A.NUM_DND
        ,A.NOME_AGENCIA
        ,A.IDT_TML
        ,A.VLR_TRANSACAO AS VALOR
        --,A.NOM_TRANSP 
        ,A.OS

    FROM
        [MERCANTIL\B042786].TC_FAT_BASE_TRANSPORTE A WITH (NOLOCK)

    WHERE
        A.DATA_MOVIMENTO BETWEEN @dta_corte AND @dta_limite
    AND
        LTRIM(RTRIM(A.NOM_TRANSP)) = LTRIM(RTRIM(@transp))
    AND
        A.COD_IMPOSTO = @imposto

    """
    df_base = pd.read_sql(query_base, con=conn)
    for c in ['VALOR_TRANSP', 'TIPO SVC', 'ADVALOREM', 'PERMANENCIA', 'VLR_EMBARQUE', 'ISS', 'ICMS', 'TOTAL', 'OBSERVAÇÃO', 'NF']:
        df_base[c] = ''
    df_base['COD_SVC'] = df_base['COD_SVC'].astype('Int64')
    df_base['DATA'] = pd.to_datetime(df_base['DATA']).dt.strftime("%d/%m/%Y")
    df_base['NOME_AGENCIA'] = df_base['NOME_AGENCIA'].str.strip()
    df_base['OS'] = df_base['OS'].astype('Int64')
    
    return df_base

def gerar_custodia(dta_inicial_padrao, resp_dta_fim, resp_transp):
    ### CUSTÓDIA
    query_custodia = f"""
    DECLARE @dta_corte AS DATE 
    DECLARE @dta_limite AS DATE 
    DECLARE @transp AS VARCHAR(50)

    SET @dta_corte = '{dta_inicial_padrao}'
    SET @dta_limite = '{resp_dta_fim}'
    SET @transp = '{resp_transp}'

    SELECT 
         A.COD_SVC
        ,A.SVC
        ,A.DATA
        ,CAST(A.COD_CEN AS VARCHAR) AS COD_CEN
        ,A.DES_CEN
        --,NULL                       AS NUM_DND
        --,NULL                       AS NOME_AGENCIA
        --,NULL                       AS IDT_TML
        ,A.SALDO_PERNOITE           AS VALOR
        --,A.NOM_TRANSP 
        --,A.OS
    FROM [MERCANTIL\B042786].TC_FAT_BASE_CUSTODIA A
    WHERE A.DATA BETWEEN @dta_corte AND @dta_limite
    AND LTRIM(RTRIM(A.NOM_TRANSP)) = LTRIM(RTRIM(@transp)) 

    ORDER BY 3 DESC
    """
    df_custodia = pd.read_sql(query_custodia, con=conn)
    for c in ['VALOR_TRANSP', 'CUSTO_PERNOITE', 'ISS', 'TOTAL', 'NF']:
        df_custodia[c] = ''
    df_custodia['COD_SVC'] = df_custodia['COD_SVC'].astype('Int64')
    df_custodia['DATA'] = pd.to_datetime(df_custodia['DATA']).dt.strftime("%d/%m/%Y")

    return df_custodia



#%%
##### COLETANDO OPÇÕES DO USUÁRIO #####
##### TIPO IMPOSTO #####
while True:
    print('''GOSTARIA DE CRIAR OS ARQUIVOS PARA QUAL IMPOSTO? Envie...
[1] 'ISS'
[2] 'ICMS'
[3]  AMBOS
''')
    try:
        resp_imposto = int(input('Opção selecionada: '))
        if resp_imposto not in (1, 2, 3):
            print('Opção inválida! Responda novamente.\n')
        else:
            break
    except:
        print('Opção inválida! Responda novamente.\n')

if resp_imposto == 1:
    nome_imposto =  ['ISS']
    temp_imposto = "'ISS'"
elif resp_imposto == 2:
    nome_imposto =  ['ICMS']
    temp_imposto = "'ICMS'"
else:
    nome_imposto = ['ISS', 'ICMS']
    temp_imposto = "'ISS', 'ICMS'"



#%%
##### MES DA DATA_INICIO DO CORTE #####
while True:
    print(f'''\nQUAL O MÊS DA DATA INICIAL DO CORTE PARA CRIAÇÃO DO ARQUIVO DE {"ISS" if resp_imposto == 1 else "ICMS"}?
[ *** ENVIE UM MÊS DE 1 A 12 *** ]
''')
    try:
        resp_mes_ini = int(input('Mês base: '))
        if resp_mes_ini not in range(1, 13):
            print('Opção inválida! Responda novamente.')
        else:
            break
    except:
        print('Opção inválida! Responda novamente.')

#%%
##### TRANSPORTADORA #####
while True:
    print('''\nGOSTARIA DE CRIAR O ARQUIVO DE QUAL TRANSPORTADORA?
[1] BRINKS
[2] PROSEGUR
[3] PROTEGE
[4] TODAS
''')
    try:
        resp_transp = int(input('Opção selecionada: '))
        if resp_transp not in (1, 2, 3, 4):
            print('Opção inválida! Responda novamente.')
        else:
            break
    except:
        print('Opção inválida! Responda novamente.')
if resp_transp == 1:
    resp_transp = ['BRINKS']
    temp_transp = "'BRINKS'"
elif resp_transp == 2:
    resp_transp = ['PROSEGUR']
    temp_transp = "'PROSEGUR'"
elif resp_transp == 3:
    resp_transp = ['PROTEGE']
    temp_transp = "'PROTEGE'"
else:
    resp_transp = ['BRINKS', 'PROSEGUR', 'PROTEGE']
    temp_transp = "'BRINKS', 'PROSEGUR', 'PROTEGE'"



# %%
##### Conexão #####
connection_string = (
    r"Driver=SQL Server Native Client 11.0;"
    r"Server=SWDVMA0199;"
    r"Database=GNU;"
    r"Trusted_Connection=yes;"
)
connection_url = URL.create(
    "mssql+pyodbc", 
    query={"odbc_connect": connection_string}
)
engine = sqlalchemy.create_engine(connection_url, fast_executemany=True, connect_args={'connect_timeout': 10}, echo=False)
conn = engine.connect()

# %%
##### Datas de corte #####
query_impostos = f"""
    DECLARE @mes_ini AS tinyint

    SET @mes_ini = {resp_mes_ini}

    SELECT  A.TRANSPORTADORA
        ,B.COD_IMPOSTO
        ,CASE WHEN CAST(CAST(DATEPART(YEAR,GETDATE()) AS VARCHAR) 
                + '-' 
                + CAST(@mes_ini AS VARCHAR) 
                + '-' 
                + CAST(A.DTA_INICIO AS VARCHAR) AS DATE) 
        > CAST(GETDATE() AS DATE) 
        THEN CAST(CAST(DATEPART(YEAR,GETDATE()) - 1 AS VARCHAR) 
                + '-' 
                + CAST(@mes_ini AS VARCHAR) 
                + '-' 
                + CAST(A.DTA_INICIO AS VARCHAR) AS DATE)  
        ELSE CAST(CAST(DATEPART(YEAR,GETDATE()) AS VARCHAR) 
                + '-' 
                + CAST(@mes_ini AS VARCHAR) 
                + '-' 
                + CAST(A.DTA_INICIO AS VARCHAR) AS DATE) 
            END AS DTA_INICIO

        ,CASE WHEN A.DTA_FIM = 31
        THEN
			CASE WHEN 
				DATEADD(DAY, -1,
					DATEADD(MONTH, 1,
						CASE WHEN CAST(CAST(DATEPART(YEAR,GETDATE()) AS VARCHAR) 
								+ '-' 
								+ CAST(@mes_ini AS VARCHAR) 
								+ '-' 
								+ CAST(A.DTA_INICIO AS VARCHAR) AS DATE) 
						> CAST(GETDATE() AS DATE) 
						THEN CAST(CAST(DATEPART(YEAR,GETDATE()) - 1 AS VARCHAR) 
								+ '-' 
								+ CAST(@mes_ini AS VARCHAR) 
								+ '-' 
								+ '01' AS DATE)  
						ELSE CAST(CAST(DATEPART(YEAR,GETDATE()) AS VARCHAR) 
								+ '-' 
								+ CAST(@mes_ini AS VARCHAR) 
								+ '-' 
								+ '01' AS DATE) 
						END)
					) >= CAST(GETDATE() AS DATE) THEN CAST(GETDATE() -1 AS DATE)
				ELSE 
					DATEADD(DAY, -1,
						DATEADD(MONTH, 1,
							CASE WHEN CAST(CAST(DATEPART(YEAR,GETDATE()) AS VARCHAR) 
									+ '-' 
									+ CAST(@mes_ini AS VARCHAR) 
									+ '-' 
									+ CAST(A.DTA_INICIO AS VARCHAR) AS DATE) 
							> CAST(GETDATE() AS DATE) 
							THEN CAST(CAST(DATEPART(YEAR,GETDATE()) - 1 AS VARCHAR) 
									+ '-' 
									+ CAST(@mes_ini AS VARCHAR) 
									+ '-' 
									+ '01' AS DATE)  
							ELSE CAST(CAST(DATEPART(YEAR,GETDATE()) AS VARCHAR) 
									+ '-' 
									+ CAST(@mes_ini AS VARCHAR) 
									+ '-' 
									+ '01' AS DATE) 
							END))
                END
            ELSE
                CASE WHEN 
                    DATEADD(MONTH, 1, 
                        CASE WHEN CAST(CAST(DATEPART(YEAR,GETDATE()) AS VARCHAR) 
                                + '-' 
                                + CAST(@mes_ini AS VARCHAR) 
                                + '-' 
                                + CAST(A.DTA_FIM AS VARCHAR) AS DATE) 
                        > CAST(GETDATE() AS DATE) 
                        THEN CAST(CAST(DATEPART(YEAR,GETDATE()) - 1 AS VARCHAR) 
                                + '-' 
                                + CAST(@mes_ini AS VARCHAR) 
                                + '-' 
                                + CAST(A.DTA_FIM AS VARCHAR) AS DATE)  
                        ELSE CAST(CAST(DATEPART(YEAR,GETDATE()) AS VARCHAR) 
                                + '-' 
                                + CAST(@mes_ini AS VARCHAR) 
                                + '-' 
                                + CAST(A.DTA_FIM AS VARCHAR) AS DATE) 
                        END
                        )
                    > CAST(GETDATE() AS DATE) 
                    THEN CAST(GETDATE() - 1 AS DATE)
                ELSE 				
                    DATEADD(MONTH, 1, 
                        CASE WHEN CAST(CAST(DATEPART(YEAR,GETDATE()) AS VARCHAR) 
                                + '-' 
                                + CAST(@mes_ini AS VARCHAR) 
                                + '-' 
                                + CAST(A.DTA_FIM AS VARCHAR) AS DATE) 
                        > CAST(GETDATE() AS DATE) 
                        THEN CAST(CAST(DATEPART(YEAR,GETDATE()) - 1 AS VARCHAR) 
                                + '-' 
                                + CAST(@mes_ini AS VARCHAR) 
                                + '-' 
                                + CAST(A.DTA_FIM AS VARCHAR) AS DATE)  
                        ELSE CAST(CAST(DATEPART(YEAR,GETDATE()) AS VARCHAR) 
                                + '-' 
                                + CAST(@mes_ini AS VARCHAR) 
                                + '-' 
                                + CAST(A.DTA_FIM AS VARCHAR) AS DATE) 
                        END
                        )
                END
            END AS DTA_FIM
    FROM [MERCANTIL\B042786].TC_FAT_DTAS_CORTE A
	LEFT JOIN [MERCANTIL\B042786].TC_FAT_IMPOSTO B
	ON A.COD_IMPOSTO = B.COD_IMPOSTO
    WHERE A.TRANSPORTADORA IN ({temp_transp})
    AND B.IMPOSTO IN ({temp_imposto})
"""

df_impostos = pd.read_sql(query_impostos, con=conn)



# %%
##### Gerar arquivo escolhido #####
lista_transportadoras = df_impostos['TRANSPORTADORA'].values.tolist()
lista_impostos = df_impostos['COD_IMPOSTO'].values.tolist()
lista_dtas_inicio = df_impostos['DTA_INICIO'].values.tolist()
lista_dtas_fim = df_impostos['DTA_FIM'].values.tolist()

# globals()[f'filiais_{transp}'] = 
for transp, imp, dta_ini, dta_f in zip(lista_transportadoras, lista_impostos, lista_dtas_inicio, lista_dtas_fim):

    base = gerar_base(
            dta_ini,
            dta_f,
            transp,
            imp
        )
    if imp == 1:
        custodia = gerar_custodia(
                    dta_ini,
                    dta_f,
                    transp
                )


    ##### Tabelas inclusão #####
    path = r"K:\GSAS\09 - Coordenacao Gestao Numerario\09 Prototipos SSIS\B042786\SISS\FATURAMENTO\V3.0"
    modelo = path + r"\MODELO PRE-FATURA.xlsx"

    incluir_base = pd.read_excel(modelo, sheet_name="INCLUIR BASE")

    legenda = pd.read_excel(modelo, sheet_name="LEGENDA")
    legenda.rename(columns={'Unnamed: 2': ''}, inplace=True)

    if imp == 1:
        incluir_custodia = pd.read_excel(modelo, sheet_name="INCLUIR CUSTODIA")


    ##### Gerar planilha #####
    today = datetime.today().date().strftime("%d-%m-%Y")
    #nome_inicio = datetime.strptime(dta_ini, "%Y-%m-%d").strftime("%d-%m-%Y")
    #nome_fim = datetime.strptime(dta_f, "%Y-%m-%d").strftime("%d-%m-%Y")
    nome_inicio = dta_ini.strftime("%d-%m-%Y")
    nome_fim = dta_f.strftime("%d-%m-%Y")


    file = path + fr"\TESTE GERA PLANS\{transp} - {'ISS' if imp == 1 else 'ICMS'} - {nome_inicio} a {nome_fim}.xlsx"

    if imp == 1:
        with pd.ExcelWriter(file) as writer:
            base.to_excel(writer, sheet_name='BASE', index=False)
            custodia.to_excel(writer, sheet_name='CUSTODIA', index=False)
            incluir_base.to_excel(writer, sheet_name='INCLUIR BASE', index=False)
            incluir_custodia.to_excel(writer, sheet_name='INCLUIR CUSTODIA', index=False)
            legenda.to_excel(writer, sheet_name='LEGENDA', index=False)
    else:
        with pd.ExcelWriter(file) as writer:
            base.to_excel(writer, sheet_name='BASE', index=False)
            incluir_base.to_excel(writer, sheet_name='INCLUIR BASE', index=False)
            legenda.to_excel(writer, sheet_name='LEGENDA', index=False)


    ##### Formatação #####    
    tip_svc = pd.read_excel(modelo, sheet_name="CONFIG", usecols="C").dropna()
    obs = pd.read_excel(modelo, sheet_name="CONFIG", usecols="A").dropna()

    tip_options = '"'+",".join(map(str, tip_svc["SERVIÇOS"].values.tolist()))+'"'
    obs_options = '"'+",".join(map(str, obs["TIPOS"].values.tolist()))+'"'

    formatar_arquivo(file, imp)

    print(f"Arquivo {transp} - {'ISS' if imp == 1 else 'ICMS'} - {nome_inicio} a {nome_fim} finalizado!")

# %%
